import pandas as pd
import xlwings as xw
from openpyxl import load_workbook
from openpyxl import Workbook

GenINDEX = 0
AmpLengthINDEX  = 1
Read_SumINDEX  = 5
Bgr_SumINDEX  = 6
AllBgr_SumINDEX  = 7
FinalScoreINDEX = 12
FinalScoreCommentINDEX = 13
FinalScoreTextINDEX = 14

def GetMaxByReadSum(Gens):
    MaxReadSum=0
    retGen = None
    
    for gen in Gens:
        if gen["Read_Sum"]>MaxReadSum:
            MaxReadSum=gen["Read_Sum"]
            retGen = gen
    return retGen

def GetMinByReadSum(Gens):
    MinReadSum=999999999
    retGen =  None
    
    for gen in Gens:
        if gen["Read_Sum"]<MinReadSum:
            MinReadSum=gen["Read_Sum"]
            retGen = gen
    return retGen

def getFinal(Gens):
    final = None
    for gen in Gens:
        if "final" in gen and gen["final"]:
            final = gen
    return final

def ShowResult(Gens, sheet, GeneStartLine):
    for gen in Gens:
        #sheet.range((gen["GeneSymbolStep"]+GeneStartLine,FinalScoreINDEX)).value = ""
        #sheet.range((gen["GeneSymbolStep"]+GeneStartLine,FinalScoreCommentINDEX)).value = ""
        #sheet.range((gen["GeneSymbolStep"]+GeneStartLine,FinalScoreTextINDEX)).value = ""
        r = gen["GeneSymbolStep"]+GeneStartLine
        if "final" in gen and gen["final"]:
            sheet.cell(column=r, row=FinalScoreINDEX, value="MAX")
            if "comment" in gen :
                sheet.cell(column=r, row=FinalScoreCommentINDEX, value=gen["comment"])
            if "finalString" in gen :
                sheet.cell(column=r, row=FinalScoreTextINDEX, value=gen["finalString"])


def CalcFinalScore(Gens, BaseGen):
    print("CalcFinalScore")
    score=[0,0,0]
    if len(Gens)==0:
        return
    for gen in Gens:
        if gen["Read_Sum"]>=200:
            gen["score"]=1
            score[0]+=1
        elif gen["Read_Sum"]>=10:
            gen["score"]=2
            score[1]+=1
        else:
            gen["score"]=3
            score[2]+=1
    #print(score)
    #print(Gens)
    #a. Single 1 score
    if score[0]==1:
        print("Single 1 score")
        for gen in Gens:
            if gen["score"]==1:
                gen["final"] = True
                gen["finalString"] = 'a.Only only one have best counts '
                gen["comment"] = ''
    #For 2 or more 1 score sets:                
    elif score[0]>1:
        print("more 1 score")
        gensScore1 = list(gen for gen in Gens if gen["score"]==1)
        gens250L = list(gen for gen in gensScore1 if  gen["AmpLength"]>=45 and  gen["AmpLength"]<=250)
        # When two or three primer sets get a 1 score (with all amplicon lengths 45-250):
        if len(gens250L)==score[0]:
            print("two or three primer sets get a 1 score (with all amplicon lengths 45-250)")
            gens1500R = list(gen for gen in gens250L if gen["Read_Sum"]<15000)
            if len(list(gens1500R))>0:
                maxGem = GetMaxByReadSum(gens1500R)
                if maxGem:
                    maxGem["final"] = True
                    maxGem["finalString"] = 'b.for those with all primer length 45-180, pick one with highest reads '
                    maxGem["comment"] = ''
            else:
                gens1500R = list(gen for gen in gens250L if gen["Read_Sum"]>=15000)
                maxGem = GetMinByReadSum(gens1500R)
                if maxGem:
                    maxGem["final"] = True
                    maxGem["finalString"] = 'b.for those with all primer length 45-180, pick one with highest reads '
                    maxGem["comment"] = 'AB'

        # for those with one or two primer length 45-180
        else:
            print("When two or three primer sets get a 1 score (with one or two amplicon lengths > 250)")
            #c. When two or three primer sets get a 1 score (with one or two amplicon lengths > 250):
            gens250G = list(gen for gen in gensScore1 if  gen["AmpLength"]>250)
            if len(gens250G)>0:
                gens1500R = list(gen for gen in gensScore1 if gen["Read_Sum"]>=15000)
                #if all sets with 1 score > 15000, take lowest, add AB (abundant) to comment column
                if len(gens1500R)==len(gensScore1):
                    maxGem = GetMinByReadSum(gens1500R)
                    maxGem["final"] = True
                    maxGem["finalString"] = 'c. When two or three primer sets get a 1 score (with one or two amplicon lengths > 250):'
                    maxGem["comment"] = 'AB'
                #-if one or more sets have highest reads < 15000, pick best from these following these rules:
                elif len(list(gens1500R))!=len(gensScore1):
                    if len(list(gens250L))>0:
                            max250GGem = GetMaxByReadSum(gens250G)
                            max250LGem = GetMaxByReadSum(gens250L)
                            if max250GGem["Read_Sum"]>max250LGem["Read_Sum"]*3:
                                max250GGem["final"] = True
                                max250GGem["finalString"] = 'c. only take set with amplicon length > 250 if it is > 3X highest read of best set with amplicon length 45-250'
                                max250GGem["comment"] = ''
                            else:
                                max250LGem["final"] = True
                                max250LGem["finalString"] = 'c. otherwise take best from 45-250 amplicon length set(s)'
                                max250LGem["comment"] = ''
                else:
                    max250GGem = GetMaxByReadSum(gens250G)
                    max250GGem["final"] = True
                    max250GGem["finalString"] = 'if all amplicon lengths > 250, take one with highest reads'
                    max250GGem["comment"] = ''

                #-for all best picks, calculate total background as sum of col G & H background values,
                # if total background is > 20% of reads for best pick, exclude and take next best choice and enter B into comment column

                finalGen = getFinal(Gens)
                if finalGen:
                    backgroundSum = finalGen["Bgr_Sum"]+finalGen["AllBgr_Sum"]
                    if(backgroundSum>0.2*finalGen["Read_Sum"]):
                        finalGen["final"] = None
                        finalGen["finalString"] = None
                        finalGen["comment"] = None

    finalGen = getFinal(Gens)
    # Next
    if not finalGen:
        #For sets with only scores of 2s and 3s:
        if score[0]==0:
            gens2Score = list(gen for gen in Gens if  gen["score"]==2)
            #-for all 2s, pick one with highest reads, mark as 2
            if score[1]>0 and score[2]==0:
                maxGem = GetMaxByReadSum(gens2Score)
                maxGem["final"] = True
                maxGem["finalString"] = 'for all 2s, pick one with highest reads, mark as 2'
                maxGem["comment"] = '2'
            elif score[1]>0 and score[2]>0:
                maxGem = GetMaxByReadSum(gens2Score)
                maxGem["final"] = True
                maxGem["finalString"] = 'for mix of 2s and 3s, pick the 2 with highest reads, mark as 2'
                maxGem["comment"] = '2'
            else:
                gens2Score = list(gen for gen in Gens if  gen["score"]==3)
                maxGem = GetMaxByReadSum(gens2Score)
                if maxGem:
                    maxGem["final"] = True
                    maxGem["finalString"] = 'for all 3s, pick one with highest reads, mark as 3'
                    maxGem["comment"] = '3'

#e. Compare to 4th set when available.
#-if 4th set > 3x the reads of the best pick based on above, then enter L into comment window (low signal) next to mark of best set as 1, 2 or 3….
    if BaseGen:
        finalGen = getFinal(Gens)
        if finalGen:
            if BaseGen["Read_Sum"]*3>finalGen["Read_Sum"]:
                   finalGen["comment"] = 'L'+(finalGen["comment"] if "comment" in  finalGen else "")


    if len(Gens)==1:
        Gens[0]["comment"]=(Gens[0]["comment"] if  "comment" in Gens[0] else "")+"4"
    return


wb = load_workbook('PrimerPerformanceforScoringV2.xlsx')

sheet = wb.active
GeneSymbol =''
GeneSymbolStep=0
GeneStartLine=0
CountGen = 0
Gens = []
BaseGen=None
EmptyLineCount=0

for r in range(2,100000):
    ro = sheet[r]

    if not ro[GenINDEX].value:
        EmptyLineCount+=1
        if EmptyLineCount>10:
            break
        else:
            continue

    EmptyLineCount=0
    Gen = ro[GenINDEX].value
    if Gen!=GeneSymbol:
        if len(Gens)>0:
            CalcFinalScore(Gens, BaseGen)
            ShowResult(Gens,sheet,GeneStartLine -1)
            print("GeneStartLine=",GeneStartLine)
        Gens = []
        BaseGen=None
        GeneSymbol = Gen
        GeneSymbolStep = 0
        GeneStartLine = r
        CountGen+=1
        #demo
        if CountGen>300:
            break

    GeneSymbolStep+=1
    AmpLength  = ro[AmpLengthINDEX].value
    Read_Sum = ro[Read_SumINDEX].value
    Bgr_Sum = ro[Bgr_SumINDEX].value
    AllBgr_Sum = ro[AllBgr_SumINDEX].value
    Gen={"GeneSymbolStep":GeneSymbolStep, "GeneSymbol":GeneSymbol, "AmpLength":AmpLength,"Read_Sum":Read_Sum,"Bgr_Sum":Bgr_Sum, "AllBgr_Sum":AllBgr_Sum}
    if Bgr_Sum==None:
        BaseGen = Gen
    else:
        Gens.append(Gen)
 
CalcFinalScore(Gens, BaseGen)
ShowResult(Gens,sheet,GeneStartLine-1 )
print("End prog")
wb.save('result.xlsx')
#